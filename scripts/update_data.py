#!/usr/bin/env python3
"""
국제선 항공통계 대시보드 — 월별 데이터 자동 업데이트 스크립트

사용법:
  1. 항공포털(https://www.airportal.go.kr/stats/transport/chartDetail3.do)에서
     김포·인천 노선별 월별 엑셀 다운로드.
  2. data/raw/ 아래 폴더 구조로 정리:
        data/raw/김포 25년/항공통계상세조회(노선) (1).xlsx ~ (12).xlsx
        data/raw/김포 26년/항공통계상세조회(노선) (1).xlsx ~ (n).xlsx
        data/raw/인천 25년/...
        data/raw/인천 26년/...
     파일명의 (n) = n월 (그대로 두면 됨)
  3. 이 스크립트 실행:
        python3 scripts/update_data.py
  4. docs/index.html, docs/china.html 자동 갱신됨.
  5. git commit + push → GitHub Pages 자동 배포.
"""

import openpyxl
import os
import re
import json
import sys
from collections import defaultdict
from pathlib import Path
import warnings
warnings.filterwarnings('ignore')

ROOT = Path(__file__).resolve().parent.parent
RAW_DIR = ROOT / "data" / "raw"
PROCESSED_DIR = ROOT / "data" / "processed"
DOCS_DIR = ROOT / "docs"

# =============================================================================
# 팀 매핑 (T&A 사업실 7개 팀)
# =============================================================================
TEAM_IATA = {
    "Europe-mid": ["FCO","CIA","CDG","ORY","BVA","PRG","VIE","BUD","ZAG","LHR","LGW","LCY","STN","SEN",
        "FLR","VCE","TSF","MXP","LIN","BGY","AMS","NCE","NAP","CTA","PMO","TPS","EDI",
        "DBV","SPU","KEF","RVN","TOS","OSL","FRA","HHN","MUC","BER","CPH","ARN","SZG",
        "BRU","CRL","DUS","HAM","ZRH","BSL","GVA","DUB","HEL","WAW","KRK","WRO","MRS","LUX","CGN","LEJ",
        "MAN","BHX","NCL","GLA","BFS","CWL"],
    "Europe-south": ["BCN","LIS","MAD","SVQ","IST","SAW","NAV","ASR","ATH","OPO","GRX","PMI",
        "RAK","JTR","LXR","CAI","DXB","DWC","AUH","DOH","RUH","JED","TLV","AYT","ESB","ADB",
        "ADD","TBS","EVN","BEY","BSR","JUB","NBO","MCT","SHJ","AGP","CMN","LOS","JIB","GYD",
        "FAO","SCQ","BIO","VLC","HRG","SSH","TUN","ALG","ORN","FEZ"],
    "America": ["LAX","LAS","SFO","JFK","EWR","LGA","ORD","SEA","MCO","YVR","YYC","YYZ","YZF",
        "CUN","MEX","NLU","CUZ","GDL","HAV","GRU","LIM","BOG","SCL","EZE","DFW","ATL","BOS","IAD",
        "IAH","DTW","SAN","PHX","DEN","MSP","PDX","MIA","PHL","SLC","YUL","MTY","HSV","OAK","ANC","ONT",
        "BNA","SDF","CVG","RFD","AZA","MEM","IND","YHZ","POA","YEG","BHM","SJC","KOA","YWG"],
    "Oceania": ["HNL","GUM","SPN","SYD","MEL","PER","BNE","OOL","AKL","ZQN","CHC","CNS","ROR","NAN"],
    "Japan": ["NRT","HND","CTS","KIX","ITM","OKA","NGO","NKM","FUK","TAK","MYJ","KCZ","TKS","TSJ","MMY",
        "KMJ","KIJ","HIJ","SDJ","FSZ","KMQ","UKB","TOY","HKD","KMI","NGS","OIT","AXT","AOJ",
        "ISG","OKJ","ASJ","KOJ","KKJ","SHI","YGJ","HSG","AKJ","UBJ","OBO","IBR","SHM"],
    "South-Eastern Asia": ["DAD","CEB","CXR","BKI","DPS","PQC","TAG","SGN","HAN","KUL","KLO","MPH","CGK","HLP",
        "BKK","DMK","UTP","CNX","HKT","ULN","UBN","PNH","REP","VTE","LPQ","BOM","DEL","BLR","MAA","HYD",
        "CMB","DAC","KTM","RGN","MDL","BWN","CRK","DVO","ILO","TGG","JHB","LGK","SUB","HUI","VCA",
        "MNL","TAS","ALA","KTI","NQZ","FRU","BSZ","CIT","DLI","ASB","BTH","MDC","HPH","PEN","SAI"],
    "China": ["PVG","SHA","PEK","PKX","TAO","HRB","DYG","DLC","CKG","XIY","LJG","XMN",
        "SZX","HKG","MFM","TPE","TSA","RMQ","KHH","SIN",
        "HGH","CAN","CTU","KMG","WUH","CSX","TSN","SHE","CGO","NKG","SJW","NNG","FOC","KWE","KWL",
        "HET","TNA","HAK","SYX","WNZ","NGB","HFE","NBS","WUS","WUX","BHY","YIH","CGD","SWA","JJN",
        "CIH","YIW","ZUH","HIA","RIZ","INC","DSN","CGQ","JHG","DLU","KHN","XUZ","YNT","WEH","LXA","YIN",
        "YNJ","TFU","YTY","JMU","MDG","TXN","DAT","TYN","HUN","ENH","WDS","BAR","MZG","YNZ"]
}

# 다공항(메트로) 도시 통합 매핑
METRO_MAP = {
    "PVG":("Shanghai","상하이"), "SHA":("Shanghai","상하이"),
    "PEK":("Beijing","베이징"), "PKX":("Beijing","베이징"),
    "TPE":("Taipei","타이베이"), "TSA":("Taipei","타이베이"),
    "NRT":("Tokyo","도쿄"), "HND":("Tokyo","도쿄"),
    "KIX":("Osaka","오사카"), "ITM":("Osaka","오사카"),
    "NGO":("Nagoya","나고야"), "NKM":("Nagoya","나고야"),
    "BKK":("Bangkok","방콕"), "DMK":("Bangkok","방콕"),
    "ULN":("Ulaanbaatar","울란바토르"), "UBN":("Ulaanbaatar","울란바토르"),
    "KLO":("Boracay","보라카이"), "MPH":("Boracay","보라카이"),
    "CGK":("Jakarta","자카르타"), "HLP":("Jakarta","자카르타"),
    "DXB":("Dubai","두바이"), "DWC":("Dubai","두바이"),
    "IST":("Istanbul","이스탄불"), "SAW":("Istanbul","이스탄불"),
    "LHR":("London","런던"), "LGW":("London","런던"), "LCY":("London","런던"), "STN":("London","런던"),
    "CDG":("Paris","파리"), "ORY":("Paris","파리"), "BVA":("Paris","파리"),
    "MXP":("Milano","밀라노"), "LIN":("Milano","밀라노"), "BGY":("Milano","밀라노"),
    "FCO":("Roma","로마"), "CIA":("Roma","로마"),
    "VCE":("Venezia","베네치아"), "TSF":("Venezia","베네치아"),
    "BRU":("Brussels","브뤼셀"), "CRL":("Brussels","브뤼셀"),
    "HHN":("Frankfurt","프랑크푸르트"), "FRA":("Frankfurt","프랑크푸르트"),
    "CTA":("Sicily","시칠리아"), "PMO":("Sicily","시칠리아"), "TPS":("Sicily","시칠리아"),
    "JFK":("New York","뉴욕"), "EWR":("New York","뉴욕"), "LGA":("New York","뉴욕"),
    "MEX":("Mexico City","멕시코시티"), "NLU":("Mexico City","멕시코시티"),
}

DOMESTIC = {"CJU","PUS","TAE","KWJ","RSU","HIN","USN","CJJ","KPO","GMP","ICN"}

# =============================================================================
# 유틸
# =============================================================================
def to_int(x):
    if x is None or x == '': return 0
    try: return int(x)
    except:
        try: return int(float(x))
        except: return 0

def iata_re(s):
    m = re.search(r'\(([A-Z]{3})\)', s or '')
    return m.group(1) if m else None

def cityname(s):
    return re.sub(r'\s*\([A-Z]{3}\)\s*$', '', s or '').strip()

iata_to_team = {}
for team, codes in TEAM_IATA.items():
    for c in codes:
        if c not in iata_to_team:
            iata_to_team[c] = team

def metro_id(code):
    return METRO_MAP[code][0] if code in METRO_MAP else code

def metro_display_name(code, fallback):
    return METRO_MAP[code][1] if code in METRO_MAP else fallback

# =============================================================================
# 1) 데이터 추출
# =============================================================================
def extract_data():
    """
    data/raw/{공항 연도} 폴더에서 전체 여객 엑셀 읽기.
    data/raw/{공항 연도 환승} 폴더가 있으면 환승여객을 차감 (실제 목적지 출국자).
    """
    folder_pattern = re.compile(r'^(김포|인천)\s*(\d{2})년(\s*환승)?$')
    iata_master = {}
    iata_month = {}
    transit_month = {}  # (year, month, code) -> {pax, gmp_pax, icn_pax}

    for folder in sorted(RAW_DIR.iterdir()):
        if not folder.is_dir(): continue
        m = folder_pattern.match(folder.name)
        if not m:
            print(f"⏭️  스킵: {folder.name} (폴더명 패턴 불일치)")
            continue
        airport_kr, year2, is_transit = m.groups()
        is_transit = bool(is_transit)
        airport = "GMP" if airport_kr == "김포" else "ICN"
        year = 2000 + int(year2)
        label = "환승" if is_transit else "전체"
        files = list(folder.iterdir())
        if not files:
            if is_transit:
                continue  # 환승 폴더는 비어 있어도 무시 (옵션)
            print(f"⚠️  {folder.name}: 파일 없음")
            continue
        print(f"📁 {folder.name} → {airport} {year} ({label})")

        for fname in sorted(folder.iterdir()):
            if not fname.suffix == '.xlsx': continue
            mm = re.search(r'\((\d+)\)', fname.name)
            if not mm: continue
            month = int(mm.group(1))
            if not 1 <= month <= 12:
                print(f"   ⚠️  {fname.name}: 월 추출 실패")
                continue
            try:
                wb = openpyxl.load_workbook(fname, data_only=True)
                ws = wb["Data"]
            except Exception as e:
                print(f"   ❌ {fname.name}: {e}")
                continue
            count = 0
            for r in range(2, ws.max_row+1):
                row = [ws.cell(row=r, column=c).value for c in range(1, 7)]
                if row[1] is None or row[1] == "전체 합계": continue
                code = iata_re(row[2])
                if not code: continue
                ko = cityname(row[2])
                if code not in iata_master:
                    iata_master[code] = {"name": ko, "team": iata_to_team.get(code, "South-Eastern Asia")}
                key = (year, month, code)
                pax = to_int(row[4])
                flights = to_int(row[3])
                cargo = to_int(row[5])
                if is_transit:
                    # 환승여객: 별도 누적 (나중에 차감)
                    if key not in transit_month:
                        transit_month[key] = {"pax":0,"gmp_pax":0,"icn_pax":0}
                    transit_month[key]["pax"] += pax
                    if airport == "GMP":
                        transit_month[key]["gmp_pax"] += pax
                    else:
                        transit_month[key]["icn_pax"] += pax
                else:
                    # 전체 여객
                    if key not in iata_month:
                        iata_month[key] = {"pax":0,"flights":0,"cargo":0,"gmp_pax":0,"icn_pax":0,"gmp_flights":0,"icn_flights":0,"transit_pax":0}
                    d = iata_month[key]
                    d["pax"] += pax
                    d["flights"] += flights
                    d["cargo"] += cargo
                    if airport == "GMP":
                        d["gmp_pax"] += pax
                        d["gmp_flights"] += flights
                    else:
                        d["icn_pax"] += pax
                        d["icn_flights"] += flights
                count += 1
            print(f"   ✅ {fname.name}: {month}월, {count}개 노선")

    # 환승여객 차감
    if transit_month:
        adjusted = 0
        for key, t in transit_month.items():
            if key in iata_month:
                iata_month[key]["transit_pax"] = t["pax"]
                # 순여객 = 전체 - 환승
                iata_month[key]["pax"] = max(0, iata_month[key]["pax"] - t["pax"])
                iata_month[key]["gmp_pax"] = max(0, iata_month[key]["gmp_pax"] - t["gmp_pax"])
                iata_month[key]["icn_pax"] = max(0, iata_month[key]["icn_pax"] - t["icn_pax"])
                adjusted += 1
        print(f"\n🔧 환승여객 차감 적용: {adjusted}개 (도시·월), 총 환승 차감 {sum(t['pax'] for t in transit_month.values()):,}명")
    else:
        print(f"\n⚠️  환승여객 데이터 없음 — 점유율 분모에 환승객 포함됨 (점유율 실제보다 낮게 나옴)")
        print(f"   해결: data/raw/김포 25년 환승/, 인천 25년 환승/ 등에 항공포털 환승여객 다운로드 필요")

    return iata_master, iata_month

# =============================================================================
# 2) 데이터 집계 (메트로 통합 + 팀별)
# =============================================================================
def aggregate(iata_master, iata_month):
    """
    iata 단위 데이터를 메트로 통합 + 도시 요약 + 팀별 집계로 변환.
    YoY는 (현재 연도) vs (전년 동기) — 데이터에 있는 최신 연도/월 자동 감지.
    """
    # 가용 연/월 자동 감지
    years = sorted(set(y for (y,_,_) in iata_month.keys()))
    if len(years) < 2:
        print("⚠️ 비교할 전년 데이터가 부족합니다. YoY 분석을 건너뜁니다.")
        prev_year = years[0] if years else None
        curr_year = years[0] if years else None
        curr_months = sorted(set(m for (y,m,_) in iata_month.keys() if y == curr_year))
    else:
        prev_year, curr_year = years[-2], years[-1]
        curr_months = sorted(set(m for (y,m,_) in iata_month.keys() if y == curr_year))
    print(f"\n📊 분석 대상: {prev_year}년 전체 vs {curr_year}년 {curr_months[0] if curr_months else '?'}~{curr_months[-1] if curr_months else '?'}월")

    n_curr = len(curr_months)

    # 메트로 통합
    metro_data = defaultdict(lambda: {"name":"","codes":[],"team":"","is_metro":False,
                                       "months_25":[0]*12, "months_26":[0]*n_curr,
                                       "gmp_25":0,"icn_25":0,"flights_25":0,
                                       "gmp_q14":0,"icn_q14":0,"flights_q14":0,
                                       "gmp_26":0,"icn_26":0,"flights_26":0})
    for code, info in iata_master.items():
        if code in DOMESTIC: continue
        mid = metro_id(code)
        md = metro_data[mid]
        if not md["name"]:
            md["name"] = metro_display_name(code, info["name"])
            md["team"] = info["team"]
        if code not in md["codes"]:
            md["codes"].append(code)
        if mid != code:
            md["is_metro"] = True
        for m in range(1, 13):
            v = iata_month.get((prev_year, m, code))
            if v:
                md["months_25"][m-1] += v["pax"]
                md["gmp_25"] += v["gmp_pax"]
                md["icn_25"] += v["icn_pax"]
                md["flights_25"] += v["flights"]
                if m <= n_curr:
                    md["gmp_q14"] += v["gmp_pax"]
                    md["icn_q14"] += v["icn_pax"]
                    md["flights_q14"] += v["flights"]
        for i, m in enumerate(curr_months):
            v = iata_month.get((curr_year, m, code))
            if v:
                md["months_26"][i] += v["pax"]
                md["gmp_26"] += v["gmp_pax"]
                md["icn_26"] += v["icn_pax"]
                md["flights_26"] += v["flights"]

    metros = []
    for mid, d in metro_data.items():
        pax_25 = sum(d["months_25"])
        pax_25_q14 = sum(d["months_25"][:n_curr])
        pax_26_q14 = sum(d["months_26"])
        if pax_25 == 0 and pax_26_q14 == 0: continue
        yoy = ((pax_26_q14-pax_25_q14)/pax_25_q14*100) if pax_25_q14>0 else None
        metros.append({
            "id": mid, "name": d["name"], "codes": sorted(d["codes"]),
            "team": d["team"], "is_metro": d["is_metro"],
            "pax_25": pax_25, "pax_25_q14": pax_25_q14, "pax_26_q14": pax_26_q14,
            "yoy": yoy, "flights_25": d["flights_25"],
            "gmp_pax_25": d["gmp_25"], "icn_pax_25": d["icn_25"],
            "months_25": d["months_25"], "months_26": d["months_26"]
        })

    # 팀별 집계
    team_agg = defaultdict(lambda: {"cities":[],"pax_25":0,"pax_25_q14":0,"pax_26_q14":0,
                                      "months_25":[0]*12,"months_26":[0]*n_curr})
    for m in metros:
        t = m["team"]
        team_agg[t]["cities"].append(m)
        team_agg[t]["pax_25"] += m["pax_25"]
        team_agg[t]["pax_25_q14"] += m["pax_25_q14"]
        team_agg[t]["pax_26_q14"] += m["pax_26_q14"]
        for i in range(12): team_agg[t]["months_25"][i] += m["months_25"][i]
        for i in range(n_curr): team_agg[t]["months_26"][i] += m["months_26"][i]

    TEAM_ORDER = ["China","Japan","South-Eastern Asia","Europe-mid","Europe-south","America","Oceania"]
    teams_out = []
    for t in TEAM_ORDER:
        if t not in team_agg: continue
        a = team_agg[t]
        cities = sorted(a["cities"], key=lambda x:-x["pax_25"])
        yoy = ((a["pax_26_q14"]-a["pax_25_q14"])/a["pax_25_q14"]*100) if a["pax_25_q14"]>0 else None
        teams_out.append({
            "team": t, "city_count": len(cities),
            "pax_25": a["pax_25"], "pax_25_q14": a["pax_25_q14"], "pax_26_q14": a["pax_26_q14"],
            "yoy": yoy, "months_25": a["months_25"], "months_26": a["months_26"],
            "cities": cities
        })

    # 월별 시계열 (공항별)
    ts_out = {f"GMP_{prev_year}":[], f"GMP_{curr_year}":[], f"ICN_{prev_year}":[], f"ICN_{curr_year}":[]}
    for y in [prev_year, curr_year]:
        max_m = 12 if y == prev_year else n_curr
        months_range = range(1, max_m+1) if y == prev_year else curr_months
        for m in months_range:
            g = i = 0
            for code in iata_master:
                if code in DOMESTIC: continue
                v = iata_month.get((y, m, code))
                if v:
                    g += v.get("gmp_pax", 0)
                    i += v.get("icn_pax", 0)
            ts_out[f"GMP_{y}"].append({"month": m, "pax": g})
            ts_out[f"ICN_{y}"].append({"month": m, "pax": i})

    # 메인 대시보드 키 호환성 (GMP_2025/2026 형식)
    ts_legacy = {
        "GMP_2025": ts_out.get(f"GMP_{prev_year}", []),
        "GMP_2026": ts_out.get(f"GMP_{curr_year}", []),
        "ICN_2025": ts_out.get(f"ICN_{prev_year}", []),
        "ICN_2026": ts_out.get(f"ICN_{curr_year}", []),
    }

    total_pax_25 = sum(t["pax_25"] for t in teams_out)
    total_pax_25_q14 = sum(t["pax_25_q14"] for t in teams_out)
    total_pax_26_q14 = sum(t["pax_26_q14"] for t in teams_out)
    all_cities = sorted(metros, key=lambda x:-x["pax_25"])

    return {
        "meta": {"prev_year": prev_year, "curr_year": curr_year, "curr_months": curr_months},
        "totals": {
            "pax_25": total_pax_25, "pax_25_q14": total_pax_25_q14, "pax_26_q14": total_pax_26_q14,
            "yoy": (total_pax_26_q14-total_pax_25_q14)/total_pax_25_q14*100 if total_pax_25_q14>0 else 0,
            "city_count": len(all_cities)
        },
        "ts": ts_legacy, "teams": teams_out, "all_cities": all_cities
    }

# =============================================================================
# 3) 중국 데이터 분리
# =============================================================================
def build_china(main_data):
    china = next((t for t in main_data["teams"] if t["team"]=="China"), None)
    if not china:
        return None

    mainland = {"PVG","SHA","PEK","PKX","TAO","HRB","DYG","DLC","CKG","XIY","LJG","XMN","SZX",
                "HGH","CAN","CTU","KMG","WUH","CSX","TSN","SHE","CGO","NKG","SJW","NNG","FOC","KWE","KWL",
                "HET","TNA","HAK","SYX","WNZ","NGB","HFE","NBS","WUS","WUX","BHY","YIH","CGD","SWA","JJN",
                "CIH","YIW","ZUH","HIA","RIZ","INC","DSN","CGQ","JHG","DLU","KHN","XUZ","YNT","WEH","LXA","YIN",
                "YNJ","TFU","YTY","JMU","MDG","TXN","DAT","TYN","ENH","WDS","BAR","YNZ"}
    hk = {"HKG"}; macau = {"MFM"}
    taiwan = {"TPE","TSA","RMQ","KHH","HUN","MZG"}; sg = {"SIN"}

    def classify(codes):
        s = set(codes)
        if s & hk: return "홍콩"
        if s & macau: return "마카오"
        if s & taiwan: return "대만"
        if s & sg: return "싱가포르"
        if s & mainland: return "중국 본토"
        return "기타"

    for c in china["cities"]:
        c["subregion"] = classify(c["codes"])

    sub_agg = defaultdict(lambda: {"cities":[],"pax_25":0,"pax_25_q14":0,"pax_26_q14":0,
                                     "months_25":[0]*12,"months_26":[0]*len(main_data["meta"]["curr_months"])})
    for c in china["cities"]:
        s = c["subregion"]
        sub_agg[s]["cities"].append(c)
        sub_agg[s]["pax_25"] += c["pax_25"]
        sub_agg[s]["pax_25_q14"] += c["pax_25_q14"]
        sub_agg[s]["pax_26_q14"] += c["pax_26_q14"]
        for i in range(12): sub_agg[s]["months_25"][i] += c["months_25"][i]
        for i in range(len(c["months_26"])): sub_agg[s]["months_26"][i] += c["months_26"][i]

    SUB_ORDER = ["중국 본토","홍콩","대만","마카오","싱가포르"]
    subregions = []
    for s in SUB_ORDER:
        if s not in sub_agg: continue
        a = sub_agg[s]
        cities = sorted(a["cities"], key=lambda x:-x["pax_25"])
        yoy = (a["pax_26_q14"]-a["pax_25_q14"])/a["pax_25_q14"]*100 if a["pax_25_q14"]>0 else None
        subregions.append({
            "name": s, "city_count": len(cities),
            "pax_25": a["pax_25"], "pax_25_q14": a["pax_25_q14"], "pax_26_q14": a["pax_26_q14"],
            "yoy": yoy, "months_25": a["months_25"], "months_26": a["months_26"], "cities": cities
        })

    return {
        "team": china, "subregions": subregions,
        "total_25": china["pax_25"], "total_yoy": china["yoy"]
    }

# =============================================================================
# 4) HTML 데이터 라인 교체
# =============================================================================
def update_html(html_path: Path, data: dict):
    """const DATA = {...}; 라인을 새 데이터로 교체"""
    if not html_path.exists():
        print(f"❌ HTML 없음: {html_path}")
        return
    text = html_path.read_text(encoding='utf-8')
    new_line = f"const DATA = {json.dumps(data, ensure_ascii=False)};"
    # const DATA = ...; 라인 찾아서 교체
    new_text = re.sub(r"const DATA = .*?;\s*$", new_line, text, count=1, flags=re.MULTILINE | re.DOTALL)
    # 위 정규식이 너무 탐욕적이면 라인 단위로
    if new_text == text:
        lines = text.split('\n')
        for i, ln in enumerate(lines):
            if ln.strip().startswith('const DATA ='):
                lines[i] = new_line
                break
        new_text = '\n'.join(lines)
    html_path.write_text(new_text, encoding='utf-8')
    print(f"✅ 업데이트: {html_path.name}")

# =============================================================================
# 메인
# =============================================================================
def main():
    if not RAW_DIR.exists():
        print(f"❌ 폴더 없음: {RAW_DIR}")
        print("data/raw/ 아래에 '김포 25년', '김포 26년', '인천 25년', '인천 26년' 폴더를 만들어 엑셀을 넣어주세요.")
        sys.exit(1)

    print("=" * 60)
    print("국제선 항공통계 대시보드 — 자동 업데이트")
    print("=" * 60)

    iata_master, iata_month = extract_data()
    print(f"\n📦 추출 완료: IATA {len(iata_master)}개, 데이터 포인트 {len(iata_month)}개")

    main_data = aggregate(iata_master, iata_month)
    china_data = build_china(main_data)

    # JSON 저장
    PROCESSED_DIR.mkdir(parents=True, exist_ok=True)
    (PROCESSED_DIR / "main.json").write_text(json.dumps(main_data, ensure_ascii=False, indent=2), encoding='utf-8')
    if china_data:
        (PROCESSED_DIR / "china.json").write_text(json.dumps(china_data, ensure_ascii=False, indent=2), encoding='utf-8')
    print(f"\n💾 JSON 저장: {PROCESSED_DIR}")

    # HTML 업데이트
    update_html(DOCS_DIR / "index.html", main_data)
    if china_data:
        update_html(DOCS_DIR / "china.html", china_data)

    # 요약 출력
    print("\n" + "=" * 60)
    print("📊 분석 결과 요약")
    print("=" * 60)
    print(f"전체 국제선: 25년 {main_data['totals']['pax_25']:,}명")
    print(f"YoY 1~{len(main_data['meta']['curr_months'])}월: {main_data['totals']['yoy']:+.1f}%")
    print(f"도시 수: {main_data['totals']['city_count']}개")
    print(f"\n팀별:")
    for t in main_data['teams']:
        yoy_str = f"{t['yoy']:+.1f}%" if t['yoy'] is not None else "N/A"
        print(f"  {t['team']:22s} {t['city_count']:>3d}개 | 25년 {t['pax_25']:>11,}명 | YoY {yoy_str}")

    print("\n다음 단계: git add . && git commit -m \"data update\" && git push")

if __name__ == "__main__":
    main()
